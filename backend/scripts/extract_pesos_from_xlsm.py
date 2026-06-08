"""Extrai pesos reais da planilha '02 - CONSULTA DOS GESTORES.xlsm'
e gera pesos_gestores.json com os valores exatos (peso positivo + deflator).
"""

import json
import openpyxl
from pathlib import Path

XLSM_PATH = Path(__file__).resolve().parents[2] / "instrucoes" / "workflow" / "02 - CONSULTA DOS GESTORES.xlsm"
OUTPUT_PATH = Path(__file__).resolve().parents[1] / "core" / "pesos_gestores.json"


def extract_pesos() -> dict:
    wb = openpyxl.load_workbook(str(XLSM_PATH), data_only=True, read_only=True)
    print(f"Abas: {wb.sheetnames}")
    
    ws = wb["Pesos"]
    
    # Read all rows
    rows = list(ws.iter_rows(values_only=True))
    
    # Print header to understand structure
    print(f"\nTotal rows: {len(rows)}")
    print(f"Header: {list(rows[0])[:15]}")
    
    # Identify columns
    header = [str(h or "").strip() for h in rows[0]]
    print(f"All headers: {header}")
    
    # Build pesos dict
    pesos = {}
    current_key = None
    
    for i, row in enumerate(rows[1:], start=2):
        row_vals = list(row)
        
        # Skip empty rows
        if not any(row_vals[:5]):
            continue
            
        # Column A = composite key (e.g. "ALERTAS PRIORITÁRIOSMotorista1")
        # Column B = alert label
        # Column C = contact type (Motorista, Cliente, etc.)
        # Column D = criterion number
        # Column E = criterion label
        # Column F = ? (possibly max points)
        # Column G = peso (positive weight)
        # Column H = ? 
        # Column I = deflator (negative weight)
        
        composite = str(row_vals[0] or "").strip()
        alert = str(row_vals[1] or "").strip()
        contact = str(row_vals[2] or "").strip()
        crit_num_raw = row_vals[3]
        crit_label = str(row_vals[4] or "").strip()
        
        if not alert or not contact or not crit_label:
            continue
            
        try:
            crit_num = int(crit_num_raw) if crit_num_raw else None
        except (ValueError, TypeError):
            continue
            
        if crit_num is None:
            continue
        
        # Find peso and deflator columns
        # Let's check all numeric values in the row
        peso = None
        deflator = None
        
        # Try columns by index (G=6, I=8 based on typical spreadsheet layout)
        for col_idx in range(5, min(len(row_vals), 15)):
            val = row_vals[col_idx]
            if isinstance(val, (int, float)):
                if peso is None and val > 0:
                    peso = val
                elif deflator is None and val < 0:
                    deflator = val
        
        if peso is None:
            peso = 0.0
        if deflator is None:
            deflator = 0.0
        
        key = f"{alert}|{contact}"
        if key not in pesos:
            pesos[key] = {
                "alert_label": alert,
                "contact_type": contact,
                "criterios": [],
            }
        
        pesos[key]["criterios"].append({
            "num": crit_num,
            "label": crit_label,
            "peso": round(float(peso), 6),
            "deflator": round(float(deflator), 6),
        })
    
    return pesos


def main():
    print(f"Lendo: {XLSM_PATH}")
    pesos = extract_pesos()
    
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(pesos, f, ensure_ascii=False, indent=2)
    
    print(f"\n✅ Gerado: {OUTPUT_PATH}")
    print(f"   {len(pesos)} combinações alerta|contato")
    for key in sorted(pesos.keys()):
        n = len(pesos[key]["criterios"])
        total_peso = sum(c["peso"] for c in pesos[key]["criterios"])
        total_defl = sum(c["deflator"] for c in pesos[key]["criterios"])
        print(f"   {key:50s} → {n:2d} critérios  Σpeso={total_peso:6.2f}  Σdeflator={total_defl:7.2f}")


if __name__ == "__main__":
    main()
