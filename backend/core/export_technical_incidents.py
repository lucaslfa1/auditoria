"""Geração da planilha Excel de incidentes técnicos de telefonia.

Papel no fluxo: consome a lista de incidentes técnicos detectados na auditoria
de telefonia (ex.: áudio de baixa qualidade, silêncio excessivo, bitrate ruim)
e produz um arquivo .xlsx formatado para download/relatório. É puro
in-memory/CPU: não acessa banco nem rede.

Sem custo de API (só openpyxl em memória; nenhuma chamada paga a Azure).
"""
import io
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def _format_sector_label(sector_id: str | None) -> str:
    labels = {
        "bas": "BAS",
        "cadastro": "Cadastro",
        "logistica_unilever": "Logística Unilever",
        "logistica": "Logística",
        "mondelez": "Mondelez",
    }
    if not sector_id:
        return "-"
    return labels.get(sector_id, sector_id)


def _format_notes(notes: list[str] | None) -> str:
    if not notes:
        return "Sem observações técnicas."
    return " | ".join(notes)


def generate_technical_incidents_excel(incidents: list[dict]) -> io.BytesIO:
    """Monta a planilha Excel de incidentes técnicos de telefonia.

    Cada item de `incidents` é um dict de incidente; as colunas são extraídas
    de chaves como `id`, `timestamp`, `operator_name`, `operator_id`,
    `sector_id`, `alert_label`, `summary` e do bloco aninhado
    `audio_quality` (com `score`, `quality`, `notes` e `details` contendo
    `silence_ratio`, `sample_rate`, `bitrate_kbps`, `duration_seconds`).
    Aplica cabeçalho estilizado, bordas, larguras de coluna, freeze pane e
    auto-filtro.

    Retorna um `io.BytesIO` posicionado no início (seek 0), pronto para envio.
    Não tem efeitos colaterais externos (tudo em memória).
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Telefonia"

    headers = [
        "ID",
        "Data/Hora",
        "Escopo",
        "Operador",
        "ID Operador",
        "Setor",
        "Alerta",
        "Score Áudio",
        "Classificação",
        "Silencio (%)",
        "Sample Rate (Hz)",
        "Bitrate (kbps)",
        "Duração (s)",
        "Notas Técnicas",
        "Resumo",
    ]

    header_fill = PatternFill(start_color="7C2D12", end_color="7C2D12", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    for row_idx, incident in enumerate(incidents, start=2):
        audio_quality = incident.get("audio_quality") or {}
        details = audio_quality.get("details") or {}
        values = [
            incident.get("id"),
            incident.get("timestamp"),
            "Telefonia",
            incident.get("operator_name") or "Não informado",
            incident.get("operator_id") or "",
            _format_sector_label(incident.get("sector_id")),
            incident.get("alert_label") or "",
            audio_quality.get("score"),
            audio_quality.get("quality") or "",
            round((details.get("silence_ratio") or 0) * 100, 2) if isinstance(details.get("silence_ratio"), (int, float)) else None,
            details.get("sample_rate"),
            details.get("bitrate_kbps"),
            details.get("duration_seconds"),
            _format_notes(audio_quality.get("notes")),
            incident.get("summary") or "",
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col_idx in (8, 10, 11, 12, 13):
                cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = {
        "A": 8,
        "B": 20,
        "C": 14,
        "D": 26,
        "E": 16,
        "F": 20,
        "G": 28,
        "H": 12,
        "I": 18,
        "J": 12,
        "K": 16,
        "L": 16,
        "M": 14,
        "N": 60,
        "O": 80,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:O{max(2, len(incidents) + 1)}"
    sheet.row_dimensions[1].height = 24

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def build_technical_incidents_filename(sector_id: str | None = None) -> str:
    """Monta o nome do arquivo de download da planilha de telefonia.

    Formato: `auditoria_telefonia[_<sector_id>]_<YYYY-MM-DD>.xlsx`, usando a
    data local de hoje. Quando `sector_id` é informado, ele entra como sufixo
    antes da data. Função pura (sem efeitos colaterais).
    """
    stamp = datetime.now().strftime("%Y-%m-%d")
    sector_part = f"_{sector_id}" if sector_id else ""
    return f"auditoria_telefonia{sector_part}_{stamp}.xlsx"
