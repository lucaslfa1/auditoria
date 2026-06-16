from repositories import audits
"""
Módulo de exportação no formato da planilha de Consulta dos Gestores.

Converte resultados de auditoria do sistema para o formato Excel usado pela gestão,
com escala 0-10, sistema SIM/NÃO e pontuação ponderada com deflators.
"""

import io
import json
import os
from datetime import datetime
from functools import lru_cache
from typing import Optional
from pathlib import Path
import logging

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from core.gestores_mapping import SECTOR_MAP, get_gestores_alert_catalog, resolve_gestores_alert

logger = logging.getLogger(__name__)

BASE_DIR = Path(__file__).resolve().parent
PESOS_JSON_PATH = BASE_DIR / "pesos_gestores.json"

# Meses em português
MESES_PT = {
    1: "Jan", 2: "Fev", 3: "Mar", 4: "Abr", 5: "Mai", 6: "Jun",
    7: "Jul", 8: "Ago", 9: "Set", 10: "Out", 11: "Nov", 12: "Dez"
}

# Cabeçalhos da aba BD (exatamente como na planilha original)
BD_HEADERS = [
    "#", "#2", "Coluna3", "Hora Atual", "Mês Ref.", "Matricula do Operador",
    "Nome do Operador", "Auditoria", "Ocorrência", "Motivo Desafio", "Setor",
    "AE/Placa", "DadosdaGravação",
    "1", "2", "3", "4", "5", "6", "7", "8", "9", "10",
    "11", "12", "13", "14", "15", "16", "17", "18",
    "nota", "Observação", "Relato da Ligação", "Modalidade da Auditoria",
    "super", "#3",
    "FIXO-1", "FIXO-2", "FIXO-3", "FIXO-4", "FIXO-5", "FIXO-6",
    "FIXO-7", "FIXO-8", "FIXO-9", "FIXO-10", "FIXO-11", "FIXO-12",
    "FIXO-13", "FIXO-14", "FIXO-15", "FIXO-16", "FIXO-17", "FIXO-18",
    "Nota Final"
]


@lru_cache(maxsize=1)
def _load_pesos_from_path(path: str) -> dict:
    pesos_path = Path(path)
    if not pesos_path.exists():
        raise FileNotFoundError(
            f"Arquivo de pesos nao encontrado: {pesos_path}. "
            "Gere o arquivo com backend/scripts/generate_pesos_gestores.py."
        )
    with open(pesos_path, "r", encoding="utf-8") as f:
        return json.load(f)


def load_pesos() -> dict:
    """Carrega os pesos da planilha dos gestores a partir do JSON extraído."""
    return _load_pesos_from_path(str(PESOS_JSON_PATH))


def find_best_pesos_key(alert_label: str, contact_type: str, pesos: dict) -> Optional[str]:
    """Encontra a melhor chave de pesos para um alerta+contato."""
    # Tentar chave exata
    key = f"{alert_label}|{contact_type}"
    if key in pesos:
        return key

    # Tentar com variações de encoding (o JSON pode ter encoding diferente)
    for k in pesos:
        k_alert, k_contact = k.split("|", 1)
        # Normalizar removendo acentos para comparação
        if _normalize(k_alert) == _normalize(alert_label) and _normalize(k_contact) == _normalize(contact_type):
            return k

    # Tentar apenas o alerta com qualquer contato
    for k in pesos:
        k_alert = k.split("|", 1)[0]
        if _normalize(k_alert) == _normalize(alert_label):
            return k

    return None


def _normalize(text: str) -> str:
    """Normaliza texto para comparação (lowercase, sem acentos comuns)."""
    import unicodedata
    text = text.lower().strip()
    # Tentar normalizar unicode
    try:
        nfkd = unicodedata.normalize("NFKD", text)
        text = "".join(c for c in nfkd if not unicodedata.combining(c))
    except Exception:
        pass
    # Remover caracteres corrompidos comuns do openpyxl
    text = text.replace("\ufffd", "")
    return text


def _normalize_binary_detail_status(value: object) -> str:
    """Reduz o status de um critério ao binário SIM/NÃO da planilha dos gestores.

    A planilha dos gestores só tem SIM/NÃO, então status não-reprovatórios
    (``pass``/``na``/``n/a``/``pending_manual``) viram ``pass`` (SIM) e os demais
    (``fail``/``partial`` e qualquer desconhecido) viram ``fail`` (NÃO) — default
    conservador.
    """
    status = str(value or "").strip().lower()
    if status in {"pass", "na", "n/a", "pending_manual"}:
        return "pass"
    if status in {"fail", "partial"}:
        return "fail"
    return "fail"


def convert_audit_to_gestores_row(
    audit_data: dict,
    row_number: int,
    pesos: dict
) -> dict:
    """
    Converte uma auditoria do nosso banco para o formato de linha da aba BD.

    audit_data: dict com campos do banco (timestamp, operator_name, operator_id,
                alert_label, sector_id, score, max_score, details_json, summary, etc.)
    row_number: número sequencial da auditoria
    pesos: dicionário de pesos carregado do JSON
    """
    details = audit_data.get("details", [])
    if isinstance(details, str):
        details = json.loads(details)

    timestamp = audit_data.get("timestamp", "")
    try:
        safe_timestamp = str(timestamp).replace("Z", "+00:00")
        dt = datetime.fromisoformat(safe_timestamp) if timestamp else datetime.now()
    except (ValueError, TypeError):
        dt = datetime.now()
    if dt.tzinfo is not None:
        dt = dt.replace(tzinfo=None)

    mes_ref = MESES_PT.get(dt.month, "Jan")
    matricula = audit_data.get("operator_id", "") or ""
    nome_operador = audit_data.get("operator_name", "") or "Não identificado"
    alert_label_raw = audit_data.get("alert_label", "") or ""
    alert_id_raw = audit_data.get("alert_id", "") or ""
    sector_raw = audit_data.get("sector_id", "") or ""
    supervisor_name = audit_data.get("supervisor", "") or ""

    alert_gestores, contact_type, resolved_alert_id = resolve_gestores_alert(
        alert_id=alert_id_raw,
        alert_label=alert_label_raw,
    )
    sector_gestores = SECTOR_MAP.get(sector_raw, sector_raw.upper())
    if not resolved_alert_id:
        logger.warning(
            "Export gestores: alerta nao mapeado dinamicamente (alert_id=%s, alert_label=%s)",
            alert_id_raw,
            alert_label_raw,
        )

    # Encontrar pesos correspondentes
    pesos_key = find_best_pesos_key(alert_gestores, contact_type, pesos)
    pesos_criterios = pesos[pesos_key]["criterios"] if pesos_key else []

    # Converter critérios: mapear status para SIM/NÃO
    criteria_responses = {}  # num -> "SIM" ou "NÃO"
    criteria_scores = {}  # num -> pontuação ponderada

    for i, detail in enumerate(details):
        status = _normalize_binary_detail_status(detail.get("status"))
        crit_num = i + 1  # Critérios são numerados sequencialmente

        if status == "pass":
            criteria_responses[crit_num] = "SIM"
            # Buscar peso correspondente
            peso_info = next((p for p in pesos_criterios if p["num"] == crit_num), None)
            if peso_info:
                criteria_scores[crit_num] = peso_info["peso"]
            else:
                criteria_scores[crit_num] = detail.get("obtainedScore", 0)
        else:
            criteria_responses[crit_num] = "NÃO"
            peso_info = next((p for p in pesos_criterios if p["num"] == crit_num), None)
            if peso_info:
                criteria_scores[crit_num] = peso_info["deflator"]
            else:
                criteria_scores[crit_num] = 0

    # Calcular nota final ponderada (soma dos FIXO)
    nota_final = sum(criteria_scores.get(i, 0) for i in range(1, 19))

    # Chave composta (formato da planilha)
    chave1 = f"{mes_ref}{matricula}{contact_type}{sector_gestores}"
    chave2 = f"{mes_ref}{matricula}{row_number}{sector_gestores}PADRÃO"

    # Montar a linha
    row = {
        "#": chave1,
        "#2": chave2,
        "Coluna3": row_number,
        "Hora Atual": dt,
        "Mês Ref.": mes_ref,
        "Matricula do Operador": matricula,
        "Nome do Operador": nome_operador,
        "Auditoria": row_number,
        "Ocorrência": alert_gestores,
        "Motivo Desafio": contact_type,
        "Setor": sector_gestores,
        "AE/Placa": "",
        "DadosdaGravação": "",
    }

    # Critérios 1-18 (SIM/NÃO)
    for i in range(1, 19):
        row[str(i)] = criteria_responses.get(i, "")

    row["nota"] = round(nota_final, 4)

    # Observação: listar critérios que falharam para contexto do gestor
    failed_criteria = []
    for i, detail in enumerate(details):
        if _normalize_binary_detail_status(detail.get("status")) == "fail":
            label = detail.get("label", f"Critério {i+1}")
            failed_criteria.append(label)
    row["Observação"] = "; ".join(failed_criteria) if failed_criteria else ""

    # Relato da Ligação: combinar resumo da IA + trecho da transcrição
    summary = audit_data.get("summary", "") or ""
    transcription_text = audit_data.get("transcription_text", "") or ""
    if transcription_text:
        # Limitar a 2000 chars para não estourar a célula
        relato = f"{summary}\n\n--- Transcrição ---\n{transcription_text[:2000]}"
    else:
        relato = summary

    feedback_data = audit_data.get("feedback") or {}
    feedback_texto = (feedback_data.get("feedback_texto") or "").strip() if isinstance(feedback_data, dict) else ""
    pontos_melhoria = (feedback_data.get("pontos_melhoria") or "").strip() if isinstance(feedback_data, dict) else ""
    gestor_nome = (feedback_data.get("gestor_nome") or "").strip() if isinstance(feedback_data, dict) else ""
    if feedback_texto or pontos_melhoria:
        header = "--- Feedback do Supervisor"
        if gestor_nome:
            header += f" ({gestor_nome})"
        header += " ---"
        relato += f"\n\n{header}"
        if feedback_texto:
            relato += f"\n{feedback_texto}"
        if pontos_melhoria:
            relato += f"\n\nPlano de ação:\n{pontos_melhoria}"

    row["Relato da Ligação"] = relato
    row["Modalidade da Auditoria"] = "PADRÃO"
    row["super"] = supervisor_name
    row["#3"] = f"{mes_ref}{matricula}"

    # FIXO-1 a FIXO-18 (pontuação ponderada por critério)
    for i in range(1, 19):
        row[f"FIXO-{i}"] = ""

    row["Nota Final"] = round(nota_final, 4)

    return row


def generate_gestores_excel(audits: list[dict]) -> io.BytesIO:
    """
    Gera um arquivo Excel no formato da planilha de Consulta dos Gestores.
    Formatação replicada da planilha original '02 - CONSULTA DOS GESTORES.xlsm'.

    audits: lista de dicionários com dados de auditorias do banco
    Retorna BytesIO com o arquivo Excel
    """
    wb = openpyxl.Workbook()
    pesos = load_pesos()

    # ── Aba BD ──
    ws = wb.active
    ws.title = "BD"

    # ── Estilos (replicando a planilha original) ──
    # Header: fundo azul escuro (#4472C4), texto branco bold, centralizado
    hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hdr_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Bordas finas cinza como na planilha original
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # Dados: fonte Calibri 11
    data_font = Font(name="Calibri", size=11)
    data_align_left = Alignment(horizontal="left", vertical="center")
    data_align_center = Alignment(horizontal="center", vertical="center")
    data_align_right = Alignment(horizontal="right", vertical="center")

    # SIM/NÃO: verde/vermelho suave (estilo Excel condicional padrão)
    sim_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    sim_font = Font(name="Calibri", size=11, bold=True, color="006100")
    nao_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    nao_font = Font(name="Calibri", size=11, bold=True, color="9C0006")

    # Linha alternada (zebra): cinza bem claro
    zebra_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # Colunas de critério (N=14 até AE=31)
    CRIT_COLS = set(range(14, 32))  # columns N through AE (criteria 1-18)
    # Colunas FIXO (AL=38 até BC=55)
    FIXO_COLS = set(range(38, 56))
    # Colunas de nota
    NOTA_COL = 32   # AF = nota
    NOTA_FINAL_COL = 56  # BD = Nota Final

    # ── Cabeçalho (row 1) ──
    ws.row_dimensions[1].height = 30
    for col_idx, header in enumerate(BD_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin_border

    # ── Dados ──
    for row_idx, audit in enumerate(audits, 1):
        row_data = convert_audit_to_gestores_row(audit, row_idx, pesos)
        excel_row = row_idx + 1
        is_zebra = row_idx % 2 == 0

        for col_idx, header in enumerate(BD_HEADERS, 1):
            value = row_data.get(header, "")
            cell = ws.cell(row=excel_row, column=col_idx, value=value if value != "" else None)
            cell.font = data_font
            cell.border = thin_border

            # Zebra striping
            if is_zebra:
                cell.fill = zebra_fill

            # ── Formatação por tipo de coluna ──

            # Critérios SIM/NÃO (colunas N-AE)
            if col_idx in CRIT_COLS:
                cell.alignment = data_align_center
                if value == "SIM":
                    cell.font = sim_font
                    cell.fill = sim_fill
                elif value == "NÃO":
                    cell.font = nao_font
                    cell.fill = nao_fill

            # Colunas FIXO: numérico com 4 casas, alinhado à direita
            elif col_idx in FIXO_COLS:
                cell.alignment = data_align_right
                cell.number_format = "0.0000"
                if isinstance(value, (int, float)) and value < 0:
                    cell.font = Font(name="Calibri", size=11, color="CC0000")

            # Nota e Nota Final
            elif col_idx in (NOTA_COL, NOTA_FINAL_COL):
                cell.alignment = data_align_center
                cell.number_format = "0.00"
                cell.font = Font(name="Calibri", size=11, bold=True)

            # Data/Hora (coluna D)
            elif col_idx == 4:
                cell.alignment = data_align_center
                cell.number_format = "DD/MM/YYYY HH:MM:SS"

            # Mês Ref, Modalidade, Setor (centralizados)
            elif col_idx in (5, 8, 11, 35):  # E, H, K, AI
                cell.alignment = data_align_center

            # Matrícula (centralizado)
            elif col_idx == 6:  # F
                cell.alignment = data_align_center

            # Nome do Operador (esquerda)
            elif col_idx == 7:  # G
                cell.alignment = data_align_left

            # Ocorrência e Motivo Desafio (esquerda)
            elif col_idx in (9, 10):  # I, J
                cell.alignment = data_align_left

            # Relato da Ligação (wrap text)
            elif col_idx == 34:  # AH
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

            # Chaves compostas e #3
            elif col_idx in (1, 2, 37):  # A, B, AK
                cell.alignment = data_align_left
                cell.font = Font(name="Calibri", size=9, color="808080")

            else:
                cell.alignment = data_align_center

    # ── Largura das colunas (replicando proporções da planilha original) ──
    COL_WIDTHS = {
        "A": 32,    # # (chave composta)
        "B": 36,    # #2 (chave2)
        "C": 7,     # Coluna3 (contador)
        "D": 20,    # Hora Atual
        "E": 8,     # Mês Ref.
        "F": 14,    # Matricula do Operador
        "G": 32,    # Nome do Operador
        "H": 10,    # Auditoria
        "I": 28,    # Ocorrência
        "J": 16,    # Motivo Desafio
        "K": 14,    # Setor
        "L": 12,    # AE/Placa
        "M": 14,    # DadosdaGravação
    }
    # Critérios 1-18 (colunas N-AE): largura estreita de 5.5
    for i in range(14, 32):
        COL_WIDTHS[get_column_letter(i)] = 5.5
    # nota (AF)
    COL_WIDTHS["AF"] = 10
    # Observação (AG)
    COL_WIDTHS["AG"] = 18
    # Relato da Ligação (AH)
    COL_WIDTHS["AH"] = 45
    # Modalidade (AI)
    COL_WIDTHS["AI"] = 16
    # super (AJ)
    COL_WIDTHS["AJ"] = 8
    # #3 (AK)
    COL_WIDTHS["AK"] = 14
    # FIXO-1 a FIXO-18 (colunas AL-BC): 10
    for i in range(38, 56):
        COL_WIDTHS[get_column_letter(i)] = 10
    # Nota Final (BD)
    COL_WIDTHS["BD"] = 12

    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # ── Congelar cabeçalho ──
    ws.freeze_panes = "A2"

    # ── Auto-filtro no cabeçalho (como na planilha original) ──
    last_col = get_column_letter(len(BD_HEADERS))
    last_row = len(audits) + 1
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    # ── Altura das linhas de dados ──
    for row_idx in range(2, last_row + 1):
        ws.row_dimensions[row_idx].height = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def get_export_config_report() -> str:
    """
    Gera relatório com as configurações atuais e alternativas possíveis.
    Útil para validação e ajuste fino.
    """
    pesos = load_pesos()

    lines = [
        "=" * 60,
        "RELATÓRIO DE CONFIGURAÇÃO - EXPORTAÇÃO GESTORES",
        "=" * 60,
        "",
        "1. MAPEAMENTO DE ALERTAS (Sistema ->Planilha):",
        "-" * 40,
    ]
    for sys_id, metadata in sorted(get_gestores_alert_catalog().items()):
        lines.append(f"  {sys_id:30s} -> {metadata['gestores_label']}")

    lines.extend([
        "",
        "2. MAPEAMENTO DE SETORES (Sistema ->Planilha):",
        "-" * 40,
    ])
    for sys_id, gestores_name in sorted(SECTOR_MAP.items()):
        lines.append(f"  {sys_id:25s} ->{gestores_name}")

    lines.extend([
        "",
        "3. TIPO DE CONTATO PADRÃO POR ALERTA:",
        "-" * 40,
    ])
    for alert_id, metadata in sorted(get_gestores_alert_catalog().items()):
        lines.append(f"  {alert_id:35s} -> {metadata['contact_type']}")

    lines.extend([
        "",
        "4. PESOS DISPONÍVEIS NA PLANILHA:",
        "-" * 40,
    ])
    for key, data in sorted(pesos.items()):
        n_criterios = len(data.get("criterios", []))
        lines.append(f"  {key:45s} ({n_criterios} critérios)")

    lines.extend([
        "",
        "5. CONFIGURACOES E FONTES:",
        "-" * 40,
        "  - Catalogo de alertas: derivado dinamicamente de scoring_rules.yaml",
        "  - Catalogo de setores: SECTOR_MAP em gestores_mapping.py",
        "  - Tipo de contato: inferido automaticamente pelo catalogo YAML",
        "  - Para alertas com multiplos contatos (Motorista/Cliente/Policia),",
        "    o tipo pode ser inferido do contexto da auditoria",
        "  - O campo 'Motivo Desafio' na planilha indica qual contato e:",
        "    Motorista, Motorista 2, Cliente, Cliente 2, Polícia, etc.",
        "  - 'Motorista 2' significa segundo contato com motorista na mesma ligação",
        "",
        "6. ESCALA DE PONTUAÇÃO:",
        "-" * 40,
        "  - Planilha usa escala 0-10",
        "  - SIM ->peso positivo (coluna G da aba Pesos)",
        "  - NÃO ->deflator negativo (coluna I da aba Pesos)",
        "  - Nota final = soma de todos os FIXO (pesos aplicados)",
        "  - Se SIM: FIXO = peso positivo",
        "  - Se NÃO: FIXO = deflator (valor negativo)",
    ])

    return "\n".join(lines)
