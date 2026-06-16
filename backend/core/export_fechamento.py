"""Exportação do Fechamento mensal no formato da planilha "Qualidade Final".

Gera o arquivo Excel consumido pelo BI a partir das linhas já consolidadas por
``core.fechamento_service.get_fechamento_rows``. O contrato (ordem/labels das colunas
em ``HEADERS``, larguras e formatação) é fixo e replica a planilha de referência — o
BI depende exatamente desse layout, então NÃO é para alterar labels nem ordem.

Colunas ``OPERACIONAL``/``TELEFÔNICA`` saem como número com 2 casas; ``PROCESSO`` e
``FINAL`` saem como percentual nativo do Excel quando numéricos (valores textuais
como "Adeus"/"Adeus" permanecem texto).

Sem custo de API: só leitura de banco (via ``get_fechamento_rows``) e geração local
do .xlsx (CPU/memória).
"""
import io
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from core.fechamento_service import get_fechamento_rows

# Ordem e largura do modelo Qualidade Final.
HEADERS = [
    ('ID', 2.85546875),
    ('MÊS', 4.42578125),
    ('MATRICULA', 11.5703125),
    ('COLABORADOR', 52.7109375),
    ('OPERACIONAL', 11.85546875),
    ('TELEFÔNICA', 10.0),
    ('DESEMPENHO', 11.42578125),
    ('STATUS', 8.0),
    ('TURNO / OPERAÇÃO', 20.42578125),
    ('SUPERVISOR', 10.28515625),
    ('SETOR', 15.85546875),
    ('PROCESSO - CADEIA DE CONTATOS', 17.42578125),
    ('FINAL', 6.0),
    ('HUAWEI', 6.7109375),
    ('WEON', 6.85546875),
]


def _parse_percent(value) -> float | None:
    """Converte '70%' para 0.70 ou '-4%' para -0.04."""
    if value is None or value == '':
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(',', '.').replace('%', '')
    try:
        return float(s) / 100.0
    except ValueError:
        return None


def generate_fechamento_excel_from_rows(rows) -> bytes:
    """Monta o .xlsx do Fechamento a partir de linhas já consolidadas.

    ``rows`` é a lista de dicts vinda de ``get_fechamento_rows`` (chaves ``id``,
    ``mes_str``, ``matricula``, ``nome``, ``operacional``, ``telefonica``,
    ``desempenho``, ``status``, ``turno``, ``supervisor``, ``setor``, ``processo``,
    ``final``, ``huawei`` e o opcional ``weon``). Aplica cabeçalho, larguras e a
    formatação numérica/percentual do contrato.

    Retorna os bytes do arquivo Excel. Sem efeitos colaterais (não toca banco/rede;
    apenas escreve em memória).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planilha1"

    headers = [header for header, _ in HEADERS]
    ws.append(headers)

    header_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    header_font = Font(name="Arial", size=8, color="FFFFFF", bold=True)
    data_font = Font(name="Arial", size=10)
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col_idx, cell in enumerate(ws[1], start=1):
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center if col_idx >= 13 else Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = HEADERS[col_idx - 1][1]

    ws.row_dimensions[1].height = 21
    ws.freeze_panes = None

    for row in rows:
        excel_row = [
            row['id'],
            row['mes_str'],
            row['matricula'],
            row['nome'],
            row['operacional'],
            row['telefonica'],
            row['desempenho'],
            row['status'],
            row['turno'],
            row['supervisor'],
            row['setor'],
            row['processo'],
            row['final'],
            row['huawei'],
            row.get('weon', ''),
        ]
        
        ws.append(excel_row)
        
        current_row_idx = ws.max_row
        for col_idx, cell in enumerate(ws[current_row_idx], start=1):
            cell.font = data_font
            cell.border = border
            if col_idx in (1, 2, 5, 6, 7, 8, 12, 13):
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif col_idx == 4:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Operacional (5) e Telefonica (6): numero com 2 casas.
            if col_idx in (5, 6) and cell.value != '':
                try:
                    cell.value = float(str(cell.value).replace(',', '.'))
                    cell.number_format = '0.00'
                except ValueError:
                    pass

            # Processo (12) e Final (13): percentual nativo do Excel.
            # Valores textuais, como "Adeus", permanecem como texto.
            if col_idx in (12, 13) and cell.value not in ('', None):
                parsed = _parse_percent(cell.value)
                if parsed is not None:
                    cell.value = parsed
                    cell.number_format = '0%'
                    cell.alignment = Alignment(horizontal='center')

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_fechamento_excel(get_connection, month: int, year: int) -> bytes:
    """Gera o .xlsx do Fechamento de um mês/ano consultando o banco.

    Abre uma conexão via ``get_connection``, busca as linhas consolidadas do mês com
    ``get_fechamento_rows(conn, month, year)`` (fecha a conexão em ``finally``) e
    delega a montagem do arquivo a ``generate_fechamento_excel_from_rows``.

    Efeito colateral: leitura no banco (e fechamento da conexão). Retorna os bytes do
    Excel. Sem custo de API.
    """
    conn = get_connection()
    try:
        rows = get_fechamento_rows(conn, month, year)
    finally:
        conn.close()

    return generate_fechamento_excel_from_rows(rows)
