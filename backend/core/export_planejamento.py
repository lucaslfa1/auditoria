"""Export consolidado mensal para o setor de Planejamento.

Replica exatamente o layout de '03 - FECHAMENTO PLANEJAMENTO.xlsx'.
Preenche apenas os campos que o sistema tem — demais ficam em branco
para o Planejamento completar manualmente (OPERACIONAL, CADEIA DE CONTATOS, etc).

NÃO calcula quartil nem inventa nenhuma regra. Dados brutos.
"""

import io
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from db.domain_constants import AUDIT_STATUS_APPROVED

MESES_PT = {
    1: "JAN", 2: "FEV", 3: "MAR", 4: "ABR", 5: "MAI", 6: "JUN",
    7: "JUL", 8: "AGO", 9: "SET", 10: "OUT", 11: "NOV", 12: "DEZ",
}

# Colunas exatas do fechamento planejamento original
HEADERS = [
    "ID",                           # A - sequencial
    "MÊS",                         # B - mês referência
    "MATRICULA",                    # C - matrícula do colaborador
    "COLABORADOR",                  # D - nome do colaborador
    "OPERACIONAL",                  # E - nota operacional (outro processo)
    "TELEFÔNICA",                   # F - nota da auditoria telefônica (sistema)
    "DESEMPENHO",                   # G - classificação (BOM, etc — outro processo)
    "STATUS",                       # H - ATIVO/INATIVO
    "TURNO / OPERAÇÃO",            # I - turno / escala
    "SUPERVISOR",                   # J - supervisor
    "SETOR",                        # K - setor
    "PROCESSO - CADEIA DE CONTATOS", # L - cadeia de contatos (outro processo)
    "FINAL",                        # M - nota final (outro processo)
    "HUAWEI",                       # N - ID Huawei
    "WEON",                         # O - ID Weon
]


def _fetch_operator_audit_scores(get_connection, month: int, year: int) -> dict[str, dict]:
    """Busca nota média por operador + dados do colaboradores (RH)."""
    date_start = f"{year:04d}-{month:02d}-01"
    date_end = f"{year + 1:04d}-01-01" if month == 12 else f"{year:04d}-{month + 1:02d}-01"
    conn = get_connection()
    try:
        c = conn.cursor()

        sql = """
            WITH audit_stats AS (
                SELECT
                    operator_name,
                    AVG(score) AS media_nota,
                    MAX(max_score) AS max_score,
                    COUNT(*) AS total_ligacoes,
                    ROUND(AVG(CASE WHEN max_score > 0 THEN (score * 1.0 / max_score) * 100 ELSE 0 END), 2) AS media_percentual
                FROM audits
                WHERE status = %s
                  AND COALESCE(audit_date, timestamp)::TIMESTAMP >= %s
                  AND COALESCE(audit_date, timestamp)::TIMESTAMP < %s
                GROUP BY operator_name
            )
            SELECT
                ast.*,
                COALESCE(col.matricula, '') AS matricula,
                COALESCE(col.supervisor, '') AS supervisor,
                COALESCE(col.setor, '') AS setor,
                COALESCE(col.escala, '') AS escala,
                COALESCE(col.status, 'ATIVO') AS status,
                COALESCE(col.id_huawei, '') AS id_huawei,
                COALESCE(col.id_weon, '') AS id_weon
            FROM audit_stats ast
            LEFT JOIN colaboradores col ON col.id = (
                SELECT id FROM colaboradores c
                WHERE c.nome = ast.operator_name LIMIT 1
            )
            ORDER BY ast.operator_name
        """
        c.execute(sql, [AUDIT_STATUS_APPROVED, date_start, date_end])
        results = {}
        for row in c.fetchall():
            results[row["operator_name"]] = dict(row)
        return results
    finally:
        conn.close()


def generate_planejamento_excel(
    get_connection,
    month: int,
    year: int,
    sector_id: Optional[str] = None,
) -> io.BytesIO:
    """Gera Excel no formato exato do fechamento planejamento.

    Coluna F (TELEFÔNICA) = nota média do operador na auditoria.
    Demais colunas externas (OPERACIONAL, CADEIA DE CONTATOS, FINAL, DESEMPENHO)
    ficam em branco para o Planejamento preencher.
    """
    operators = _fetch_operator_audit_scores(get_connection, month, year)

    wb = Workbook()
    ws = wb.active
    ws.title = "Plan1"

    # Estilos (replicando o padrão da planilha original)
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill("solid", fgColor="D9E2F3")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Headers
    for col_idx, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Dados
    mes_ref = MESES_PT.get(month, str(month))
    row_num = 2
    seq = 1

    for name in sorted(operators.keys()):
        op = operators[name]

        # Filtrar por setor se especificado
        if sector_id and op.get("setor", "").lower() != sector_id.lower():
            continue

        row_data = [
            seq,                                    # A: ID
            mes_ref,                                # B: MÊS
            op.get("matricula", ""),                # C: MATRICULA
            name,                                   # D: COLABORADOR
            "",                                     # E: OPERACIONAL (outro processo)
            round(op.get("media_nota", 0), 2),      # F: TELEFÔNICA (do sistema)
            "",                                     # G: DESEMPENHO (outro processo)
            op.get("status", "ATIVO"),              # H: STATUS
            op.get("escala", ""),                    # I: TURNO / OPERAÇÃO
            op.get("supervisor", ""),                # J: SUPERVISOR
            op.get("setor", ""),                     # K: SETOR
            "",                                     # L: PROCESSO - CADEIA DE CONTATOS (outro processo)
            "",                                     # M: FINAL (outro processo)
            op.get("id_huawei", ""),                # N: HUAWEI
            op.get("id_weon", ""),                   # O: WEON
        ]

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx in (1, 2, 6):  # ID, MÊS, nota — centralizar
                cell.alignment = Alignment(horizontal="center")

        seq += 1
        row_num += 1

    # Larguras das colunas (replicando proporções da planilha original)
    widths = {
        "A": 5, "B": 6, "C": 12, "D": 35, "E": 13,
        "F": 12, "G": 13, "H": 10, "I": 18, "J": 25,
        "K": 15, "L": 28, "M": 8, "N": 8, "O": 8,
    }
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
