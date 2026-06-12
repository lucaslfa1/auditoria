import inspect
import os
import sys
import unittest
from io import BytesIO

import openpyxl

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from core.export_fechamento import HEADERS, generate_fechamento_excel
from core.fechamento_service import (
    _load_layout_seed,
    _is_receptive,
    _is_uti,
    _is_uti_rj,
    _processo_uti,
    _processo_uti_rj,
    add_fechamento_layout_operador,
    get_fechamento_rows,
    remove_fechamento_layout_operador,
    save_fechamento_overrides,
)
from db.domain_constants import (
    AUDIT_STATUS_AWAITING_PAIR,
    AUDIT_STATUS_CONTESTATION_ACCEPTED,
    AUDIT_STATUS_DISCARDED,
    AUDIT_STATUS_PENDING_APPROVAL,
    FECHAMENTO_NOTA_STATUSES,
)
from db.runtime_schema import ensure_runtime_schema


class _FakeCursor:
    def __init__(self, rows=None):
        self.rows = rows or []
        self.sql = ""
        self.params = None
        self.executions = []

    def execute(self, sql, params=None):
        self.sql = sql
        self.params = params
        self.executions.append((sql, params))

    def fetchall(self):
        return self.rows


class _LayoutModeCountCursor(_FakeCursor):
    def fetchone(self):
        return {"total": 1}


class _ScriptedCursor(_FakeCursor):
    """Cursor fake com respostas de fetchone() roteirizadas (FIFO)."""

    def __init__(self, fetchones=None, rows=None):
        super().__init__(rows=rows)
        self.fetchones = list(fetchones or [])

    def fetchone(self):
        if not self.fetchones:
            raise AssertionError("fetchone() chamado mais vezes que o roteiro previa")
        return self.fetchones.pop(0)


class _FakeConnection:
    def __init__(self, cursors):
        self._cursors = list(cursors)
        self.committed = False
        self.closed = False

    def cursor(self):
        if not self._cursors:
            raise AssertionError("cursor() called more times than expected")
        return self._cursors.pop(0)

    def commit(self):
        self.committed = True

    def close(self):
        self.closed = True


def _db_row(**overrides):
    row = {
        "colab_id": 10,
        "nome": "Ana Souza",
        "matricula": "123",
        "supervisor": "Sup A",
        "setor": "uti",
        "escala": "Verde",
        "id_huawei": "HW-1",
        "status": "ATIVO",
        "auditavel": 1,
        "nota_mot": 0,
        "nota_pa": 0,
        "nota_cli": 0,
        "nota_policia": 0,
        "operacional_override": None,
        "telefonica_override": None,
        "desempenho_override": None,
        "processo_override": None,
        "final_override": None,
        "media_auditoria": 8.0,
    }
    row.update(overrides)
    return row


class TestFechamentoModule(unittest.TestCase):
    def test_rows_use_audit_date_range_for_month(self):
        cursor = _FakeCursor(rows=[_db_row()])
        conn = _FakeConnection([cursor])

        rows = get_fechamento_rows(conn, 4, 2026)

        self.assertEqual(rows[0]["operacional"], "8.0")
        self.assertIn("COALESCE(audit_date, timestamp)::TIMESTAMP >= %s", cursor.sql)
        self.assertIn("COALESCE(audit_date, timestamp)::TIMESTAMP < %s", cursor.sql)
        self.assertIn("NULLIF(c.supervisor, '')", cursor.sql)
        self.assertNotIn("EXTRACT(MONTH", cursor.sql)
        self.assertEqual(cursor.params, (list(FECHAMENTO_NOTA_STATUSES), "2026-04-01", "2026-05-01", 4, 2026))

    def test_media_conta_toda_auditoria_salva_nao_so_aprovada(self):
        # Revisao 2026-06-12 (item 1): a nota sumia do fechamento porque a
        # media so contava status 'approved' (dependia do supervisor aprovar
        # no portal). Toda auditoria salva deve valer nota; descartes e
        # contestacoes aceitas continuam fora.
        cursor = _FakeCursor(rows=[_db_row()])
        conn = _FakeConnection([cursor])

        get_fechamento_rows(conn, 4, 2026)

        self.assertIn("WHERE status = ANY(%s)", cursor.sql)
        statuses = cursor.params[0]
        self.assertIsInstance(statuses, list)  # tuple viraria "record" no psycopg2
        self.assertIn(AUDIT_STATUS_AWAITING_PAIR, statuses)
        self.assertIn(AUDIT_STATUS_PENDING_APPROVAL, statuses)
        self.assertNotIn(AUDIT_STATUS_DISCARDED, statuses)
        self.assertNotIn(AUDIT_STATUS_CONTESTATION_ACCEPTED, statuses)

    def test_rows_exclude_removed_operations_and_nameless_telephony_services(self):
        cursor = _FakeCursor(rows=[
            _db_row(colab_id=1, nome="Operador Sanofi", escala="Time Sanofi"),
            _db_row(
                colab_id=2,
                nome="",
                escala="Telefonia",
                telefonia_account="servico.telefonia",
                organizacao_telefonia="LOGISTICA",
                id_telefonia="999",
            ),
            _db_row(colab_id=3, nome="Operador Valido", escala="LOGISTICA"),
        ])
        conn = _FakeConnection([cursor])

        rows = get_fechamento_rows(conn, 4, 2026)

        self.assertEqual([row["nome"] for row in rows], ["Operador Valido"])
        self.assertEqual(rows[0]["id"], 1)

    def test_desempenho_uses_final_audit_score_after_review(self):
        cursor = _FakeCursor(rows=[
            _db_row(colab_id=1, nome="Ana Alta", media_auditoria=8.0),
            _db_row(colab_id=2, nome="Bia Baixa", media_auditoria=7.99),
            _db_row(colab_id=3, nome="Carla Sem Nota", media_auditoria=None),
        ])
        conn = _FakeConnection([cursor])

        rows = get_fechamento_rows(conn, 4, 2026)

        desempenho_by_nome = {row["nome"]: row["desempenho"] for row in rows}
        self.assertEqual(desempenho_by_nome["Ana Alta"], "BOM")
        self.assertEqual(desempenho_by_nome["Bia Baixa"], "RUIM")
        self.assertEqual(desempenho_by_nome["Carla Sem Nota"], "")

    def test_nao_auditavel_keeps_ativo_status_in_fechamento(self):
        # BUG-026 (2026-05-26): operadores ATIVO+auditavel=0 (ex: Fernanda Sant
        # Anna) devem aparecer no fechamento como ATIVO. A flag `auditavel` so
        # vale para o pipeline da IA; o fechamento exibe o status real do
        # colaborador. Anteriormente _resolve_fechamento_status forcava
        # INATIVO e escondia esses operadores do relatorio mensal.
        cursor = _FakeCursor(rows=[
            _db_row(status="ATIVO", auditavel=0, media_auditoria=8.5),
        ])
        conn = _FakeConnection([cursor])

        rows = get_fechamento_rows(conn, 4, 2026)

        self.assertEqual(rows[0]["status"], "ATIVO")
        # desempenho continua sendo BOM porque media >= threshold (8.5 >= 7)
        self.assertEqual(rows[0]["desempenho"], "BOM")

    def test_save_only_persists_overrides_that_differ_from_base_row(self):
        insert_cursor = _FakeCursor()
        select_cursor = _FakeCursor(rows=[_db_row()])
        conn = _FakeConnection([insert_cursor, select_cursor])

        save_fechamento_overrides(
            conn,
            4,
            2026,
            [
                {
                    "colab_id": 10,
                    "matricula": "123",
                    "nome": "Ana Souza",
                    "operacional": "8.0",
                    "telefonica": "",
                    "desempenho": "BOM",
                    "status": "ATIVO",
                    "turno": "Verde",
                    "supervisor": "Sup B",
                    "setor": "uti",
                    "nota_mot": 0,
                    "nota_pa": 0,
                    "nota_cli": 0,
                    "nota_policia": 0,
                    "processo": "70%",
                    "final": "-4%",
                    "huawei": "HW-1",
                }
            ],
        )

        saved = insert_cursor.params
        self.assertTrue(conn.committed)
        self.assertIsNone(saved["matricula"])
        self.assertIsNone(saved["nome"])
        self.assertIsNone(saved["operacional"])
        self.assertEqual(saved["supervisor"], "Sup B")

    def test_runtime_schema_creates_colaboradores_before_fechamento_fk(self):
        source = inspect.getsource(ensure_runtime_schema)

        colaboradores_pos = source.index("CREATE TABLE IF NOT EXISTS colaboradores")
        fechamento_pos = source.index("CREATE TABLE IF NOT EXISTS fechamento_cadeia_contatos")
        layout_pos = source.index("CREATE TABLE IF NOT EXISTS fechamento_layout_operadores")
        self.assertLess(colaboradores_pos, fechamento_pos)
        self.assertLess(colaboradores_pos, layout_pos)

    def test_qualidade_final_layout_seed_preserves_contract_rows(self):
        rows = _load_layout_seed()

        self.assertEqual(len(rows), 224)
        self.assertEqual(rows[0]["id_visual"], 1)
        self.assertEqual(rows[0]["nome"], "JULIANE CRISTINA DA SILVA ASSIS COLAÇO")
        self.assertEqual(rows[0]["turno"], "CADASTRO")
        self.assertEqual(rows[-1]["id_visual"], 17)
        self.assertEqual(rows[-1]["turno"], "CENTRAL - VERDE")
        mondelez_rows = [row for row in rows if row["turno"] == "MONDELEZ"]
        self.assertEqual(len(mondelez_rows), 3)
        self.assertTrue(all(row["huawei"] == "-" and row["weon"] == "-" for row in mondelez_rows))

    def test_layout_mode_hides_orphans_and_prefers_live_colaborador_data(self):
        # Revisao 2026-06-12 (item 2): colaborador apagado (FK ON DELETE SET
        # NULL) deixava a linha do layout viva com o nome congelado da
        # planilha de fevereiro. A query agora filtra orfas (c.id IS NOT
        # NULL) e usa nome/matricula/status vivos de `colaboradores`.
        layout_cursor = _FakeCursor(rows=[{
            "layout_id": 7,
            "id_visual": 3,
            "sequencia_bloco": 1,
            "posicao": 3,
            "layout_matricula": "123",
            "layout_nome": "NOME ANTIGO DA PLANILHA",
            "layout_turno": "CADASTRO",
            "layout_supervisor": "Sup A",
            "layout_setor": "CADASTRO",
            "nota_coluna": "OPERACIONAL",
            "status_base": "ATIVO",
            "layout_huawei": "HW-OLD",
            "layout_weon": "W-OLD",
            "colab_id": 10,
            "db_nome": "Nome Atual",
            "db_matricula": "456",
            "db_status": "INATIVO",
            "db_huawei": "HW-1",
            "db_weon": "W-1",
            "db_setor": "CADASTRO",
            "db_escala": "CADASTRO",
            "auditavel": 1,
            "nota_mot": 0,
            "nota_pa": 0,
            "nota_cli": 0,
            "nota_policia": 0,
            "media_auditoria": None,
        }])
        conn = _FakeConnection([
            _LayoutModeCountCursor(),
            layout_cursor,
            _FakeCursor(rows=[]),
        ])

        rows = get_fechamento_rows(conn, 4, 2026)

        self.assertIn("c.id IS NOT NULL", layout_cursor.sql)
        self.assertEqual(rows[0]["nome"], "Nome Atual")
        self.assertEqual(rows[0]["matricula"], "456")
        self.assertEqual(rows[0]["status"], "INATIVO")

    def test_layout_relink_recovers_rows_without_matricula_by_nome(self):
        # O re-vinculo a cada carga tem fallback por nome para linhas sem
        # matricula; sem ele a linha ficaria orfa para sempre (e oculta).
        count_cursor = _LayoutModeCountCursor()
        conn = _FakeConnection([
            count_cursor,
            _FakeCursor(rows=[]),
            _FakeCursor(rows=[]),
        ])

        get_fechamento_rows(conn, 4, 2026)

        relink_sqls = [sql for sql, _ in count_cursor.executions if "UPDATE fechamento_layout_operadores" in sql]
        self.assertEqual(len(relink_sqls), 2)
        self.assertIn("TRIM(c.matricula) = TRIM(l.matricula)", relink_sqls[0])
        self.assertIn("LOWER(TRIM(c.nome)) = LOWER(TRIM(l.nome))", relink_sqls[1])

    def test_layout_mode_appends_uti_collaborators_outside_fixed_layout(self):
        conn = _FakeConnection([
            _LayoutModeCountCursor(),
            _FakeCursor(rows=[]),
            _FakeCursor(rows=[
                _db_row(
                    colab_id=40,
                    nome="Operadora UTI",
                    matricula="400",
                    setor="UTI",
                    escala="Verde",
                    media_auditoria=None,
                ),
                _db_row(
                    colab_id=41,
                    nome="Operadora BAS",
                    matricula="401",
                    setor="BAS",
                    escala="Azul",
                    media_auditoria=None,
                ),
            ]),
        ])

        rows = get_fechamento_rows(conn, 4, 2026)

        self.assertTrue(conn.committed)
        self.assertEqual([row["nome"] for row in rows], ["Operadora UTI"])
        self.assertIsNone(rows[0]["layout_id"])
        self.assertEqual(rows[0]["colab_id"], 40)
        self.assertEqual(rows[0]["setor"], "UTI")
        self.assertEqual(rows[0]["processo"], "70%")

    def test_add_operador_cria_linha_nova_no_layout(self):
        # Revisao 2026-06-12 (item 3): admin/auditor pode incluir operador na
        # planilha pela UI. Sem linha previa -> INSERT no fim (bloco novo).
        colab = {
            "id": 50,
            "nome": "Operadora Nova",
            "matricula": "500",
            "supervisor": "Sup B",
            "setor": "UTI",
            "escala": "Verde",
            "status": "ATIVO",
            "id_huawei": "HW-50",
            "id_weon": "W-50",
        }
        cursor = _ScriptedCursor(fetchones=[colab, None, {"max_seq": 9}, {"id": 99}])
        conn = _FakeConnection([cursor])

        result = add_fechamento_layout_operador(conn, 50)

        self.assertEqual(result, {"layout_id": 99, "reativado": False})
        self.assertTrue(conn.committed)
        insert_sql = cursor.executions[-1][0]
        self.assertIn("INSERT INTO fechamento_layout_operadores", insert_sql)
        self.assertIn("RETURNING id", insert_sql)
        insert_params = cursor.executions[-1][1]
        self.assertEqual(insert_params[0], 10)  # sequencia_bloco = max + 1

    def test_add_operador_reativa_linha_existente(self):
        colab = {"id": 50, "nome": "Operadora", "matricula": "500", "supervisor": "",
                 "setor": "UTI", "escala": "Verde", "status": "ATIVO",
                 "id_huawei": "", "id_weon": ""}
        cursor = _ScriptedCursor(fetchones=[colab, {"id": 7}])
        conn = _FakeConnection([cursor])

        result = add_fechamento_layout_operador(conn, 50)

        self.assertEqual(result, {"layout_id": 7, "reativado": True})
        self.assertIn("SET ativo = TRUE", cursor.executions[-1][0])

    def test_remove_operador_desativa_linha_do_layout(self):
        cursor = _ScriptedCursor()
        conn = _FakeConnection([cursor])

        remove_fechamento_layout_operador(conn, layout_id=7)

        self.assertTrue(conn.committed)
        self.assertIn("SET ativo = FALSE", cursor.executions[-1][0])

    def test_remove_operador_dinamico_materializa_linha_desativada(self):
        # Linha extra-UTI (sem layout_id): remover cria linha desativada no
        # layout, e o NOT EXISTS do appender impede que ela volte.
        colab = {"id": 60, "nome": "Operadora UTI", "matricula": "600",
                 "supervisor": "Sup C", "setor": "UTI", "escala": "Verde",
                 "status": "ATIVO"}
        cursor = _ScriptedCursor(fetchones=[None, colab, {"max_seq": 12}])
        conn = _FakeConnection([cursor])

        remove_fechamento_layout_operador(conn, colaborador_id=60)

        self.assertTrue(conn.committed)
        insert_sql = cursor.executions[-1][0]
        self.assertIn("INSERT INTO fechamento_layout_operadores", insert_sql)
        self.assertIn("FALSE", insert_sql)

    def test_extra_uti_appender_suprime_colaborador_com_linha_no_layout(self):
        extra_cursor = _FakeCursor(rows=[])
        conn = _FakeConnection([
            _LayoutModeCountCursor(),
            _FakeCursor(rows=[]),
            extra_cursor,
        ])

        get_fechamento_rows(conn, 4, 2026)

        self.assertIn("NOT EXISTS", extra_cursor.sql)
        self.assertIn("lx.colaborador_id = c.id", extra_cursor.sql)

    def test_save_persiste_id_visual_editado_em_linha_do_layout(self):
        # Coluna ID editavel (item 3): mudou na tela -> UPDATE no layout.
        layout_base_row = {
            "layout_id": 7,
            "id_visual": 3,
            "layout_matricula": "123",
            "layout_nome": "Operadora",
            "layout_turno": "CADASTRO",
            "layout_supervisor": "Sup A",
            "layout_setor": "CADASTRO",
            "nota_coluna": "OPERACIONAL",
            "status_base": "ATIVO",
            "layout_huawei": "",
            "layout_weon": "",
            "colab_id": 10,
            "db_nome": "Operadora",
            "db_matricula": "123",
            "db_status": "ATIVO",
            "db_huawei": "",
            "db_weon": "",
            "auditavel": 1,
            "nota_mot": 0,
            "nota_pa": 0,
            "nota_cli": 0,
            "nota_policia": 0,
            "media_auditoria": None,
        }
        write_cursor = _FakeCursor()
        conn = _FakeConnection([
            write_cursor,
            _LayoutModeCountCursor(),
            _FakeCursor(rows=[layout_base_row]),
            _FakeCursor(rows=[]),
        ])

        save_fechamento_overrides(conn, 4, 2026, [{
            "layout_id": 7,
            "colab_id": 10,
            "id": 12,
            "matricula": "123",
            "nome": "Operadora",
            "operacional": "",
            "telefonica": "",
            "desempenho": "",
            "status": "ATIVO",
            "turno": "CADASTRO",
            "supervisor": "Sup A",
            "setor": "CADASTRO",
            "nota_mot": 0,
            "nota_pa": 0,
            "nota_cli": 0,
            "nota_policia": 0,
            "processo": "70%",
            "final": "-4%",
            "huawei": "",
            "weon": "",
        }])

        id_updates = [sql for sql, _ in write_cursor.executions if "SET id_visual" in sql]
        self.assertEqual(len(id_updates), 1)

    def test_uti_rj_detection_distinguishes_from_plain_uti(self):
        self.assertTrue(_is_uti_rj("UTI RJ", ""))
        self.assertTrue(_is_uti_rj("uti-rj", ""))
        self.assertTrue(_is_uti_rj("UTI/RJ", ""))
        self.assertTrue(_is_uti_rj("", "RJ"))
        self.assertFalse(_is_uti_rj("UTI", "Verde"))
        self.assertFalse(_is_uti_rj("UTI-COMBO", ""))
        self.assertTrue(_is_uti("UTI", ""))

    def test_processo_uti_table_matches_doc(self):
        self.assertAlmostEqual(_processo_uti(4), 1.10)
        self.assertAlmostEqual(_processo_uti(3), 1.00)
        self.assertAlmostEqual(_processo_uti(2), 0.90)
        self.assertAlmostEqual(_processo_uti(1), 0.80)
        self.assertAlmostEqual(_processo_uti(0), 0.70)

    def test_processo_uti_rj_uses_official_excel_formula_after_weighted_sum(self):
        # UTI/RJ faz parte da mesma tabela de cadeia. Apenas os criterios
        # de pontuacao mudam; depois disso vale a mesma formula do DOCX.
        self.assertAlmostEqual(_processo_uti_rj(5.5), 1.10)
        self.assertAlmostEqual(_processo_uti_rj(4.5), 1.10)
        self.assertAlmostEqual(_processo_uti_rj(4.0), 1.10)
        self.assertAlmostEqual(_processo_uti_rj(3.0), 1.00)
        self.assertAlmostEqual(_processo_uti_rj(2.5), 0.90)
        self.assertAlmostEqual(_processo_uti_rj(1.5), 0.70)
        self.assertAlmostEqual(_processo_uti_rj(1.0), 0.80)
        self.assertAlmostEqual(_processo_uti_rj(0), 0.70)

    def test_uti_rj_row_applies_weighted_sum_with_official_formula(self):
        row = _db_row(
            setor="UTI RJ",
            escala="RJ",
            nota_mot=1.5,
            nota_pa=1,
            nota_cli=1.5,
            nota_policia=1.5,
        )
        cursor = _FakeCursor(rows=[row])
        conn = _FakeConnection([cursor])

        rows = get_fechamento_rows(conn, 4, 2026)

        self.assertEqual(rows[0]["processo"], "110%")
        self.assertEqual(rows[0]["final"], "4%")

    def test_uti_simples_row_applies_uti_table(self):
        row = _db_row(setor="UTI", escala="Verde", nota_mot=1, nota_pa=1, nota_cli=1, nota_policia=0)
        cursor = _FakeCursor(rows=[row])
        conn = _FakeConnection([cursor])

        rows = get_fechamento_rows(conn, 4, 2026)

        self.assertEqual(rows[0]["processo"], "100%")
        self.assertEqual(rows[0]["final"], "2%")

    def test_final_formula_matches_official_excel_formula_for_80_percent(self):
        row = _db_row(setor="UTI", escala="Verde", nota_mot=1, nota_pa=0, nota_cli=0, nota_policia=0)
        cursor = _FakeCursor(rows=[row])
        conn = _FakeConnection([cursor])

        rows = get_fechamento_rows(conn, 4, 2026)

        self.assertEqual(rows[0]["processo"], "80%")
        self.assertEqual(rows[0]["final"], "")

    def test_uti_rj_uses_same_process_formula_as_document(self):
        row = _db_row(setor="UTI RJ", escala="RJ", nota_mot=1.5, nota_pa=0, nota_cli=1.5, nota_policia=0)
        cursor = _FakeCursor(rows=[row])
        conn = _FakeConnection([cursor])

        rows = get_fechamento_rows(conn, 4, 2026)

        self.assertEqual(rows[0]["processo"], "100%")
        self.assertEqual(rows[0]["final"], "2%")

    def test_receptive_detection_handles_accents_and_variants(self):
        self.assertFalse(_is_receptive("Cadastro", ""))
        self.assertFalse(_is_receptive("Checklist", ""))
        self.assertFalse(_is_receptive("Célula", ""))
        self.assertFalse(_is_receptive("celula", ""))
        self.assertFalse(_is_receptive("celula_atendimento", ""))
        self.assertTrue(_is_receptive("", "MONDELEZ"))
        self.assertFalse(_is_receptive("UTI", "Verde"))
        self.assertFalse(_is_receptive("Distribuição", ""))
        self.assertFalse(_is_receptive("BAS", ""))

    def test_mondelez_receptive_goes_to_telefonica_column(self):
        row = _db_row(setor="Mondelez", escala="MONDELEZ-DIURNO", media_auditoria=8.5)
        cursor = _FakeCursor(rows=[row])
        conn = _FakeConnection([cursor])

        rows = get_fechamento_rows(conn, 4, 2026)

        self.assertEqual(rows[0]["telefonica"], "8.5")
        self.assertEqual(rows[0]["operacional"], "")
        self.assertEqual(rows[0]["huawei"], "-")

    def test_celula_with_accent_goes_to_operacional_column(self):
        row = _db_row(setor="Célula", escala="", media_auditoria=7.0)
        cursor = _FakeCursor(rows=[row])
        conn = _FakeConnection([cursor])

        rows = get_fechamento_rows(conn, 4, 2026)

        self.assertEqual(rows[0]["operacional"], "7.0")
        self.assertEqual(rows[0]["telefonica"], "")

    def test_export_contract_formats_excel_as_official_sheet(self):
        conn = _FakeConnection([
            _FakeCursor(rows=[
                _db_row(
                    setor="UTI RJ",
                    escala="RJ",
                    nota_mot=1.5,
                    nota_pa=1,
                    nota_cli=1.5,
                    nota_policia=1.5,
                    media_auditoria=8.25,
                )
            ])
        ])

        workbook_bytes = generate_fechamento_excel(lambda: conn, 4, 2026)
        wb = openpyxl.load_workbook(BytesIO(workbook_bytes), data_only=True)
        ws = wb["Planilha1"]

        self.assertTrue(conn.closed)
        self.assertEqual([cell.value for cell in ws[1]], [header for header, _ in HEADERS])
        self.assertIsNone(ws.freeze_panes)
        self.assertEqual(ws.max_column, 15)
        self.assertEqual(ws["E2"].value, 8.25)
        self.assertEqual(ws["E2"].number_format, "0.00")
        self.assertEqual(ws["L2"].value, 1.10)
        self.assertEqual(ws["L2"].number_format, "0%")
        self.assertEqual(ws["M2"].value, 0.04)
        self.assertEqual(ws["M2"].number_format, "0%")


if __name__ == "__main__":
    unittest.main()
