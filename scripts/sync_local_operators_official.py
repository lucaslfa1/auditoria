import argparse
import json
import os
import re
import unicodedata
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse

import openpyxl
import psycopg2
from psycopg2.extras import RealDictCursor


ROOT = Path(__file__).resolve().parents[1]
EXCEL_PATH = ROOT / "docs" / "Lista - Operadores e Supervisores.xlsx"
DEFAULT_DATABASE_URL = "postgresql://postgres:postgres@localhost:5432/auditoria"
MANUAL_OPERATOR_NAMES = {
    "BRUNA CARDOSO MONTE",
    "CINTIA CRISTINA DOMINGOS RIBEIRO",
    "GABRIELA DIEGO BUSH",
    "NATALI NEIVA DA SILVA",
    "JEAN CARLOS CONTANTINO MIRANDA",
}


def load_env_database_url() -> str:
    env_path = ROOT / "backend" / ".env"
    if env_path.exists():
        for line in env_path.read_text(encoding="utf-8", errors="ignore").splitlines():
            if line.strip().startswith("DATABASE_URL="):
                return line.split("=", 1)[1].strip().strip('"').strip("'")
    return os.environ.get("DATABASE_URL", DEFAULT_DATABASE_URL)


def normalize_text(value: object) -> str:
    normalized = unicodedata.normalize("NFD", str(value or "").strip().lower())
    without_accents = "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn")
    return re.sub(r"\s+", " ", without_accents)


def display_text(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def clean_id(value: object) -> str:
    text = display_text(value)
    if text.endswith(".0"):
        text = text[:-2]
    return "" if text == "-" else text


def connect(database_url: str):
    parsed = urlparse(database_url)
    host = (parsed.hostname or "").lower()
    database = (parsed.path or "").lstrip("/")
    if host not in {"localhost", "127.0.0.1", "::1"} or database != "auditoria":
        raise SystemExit(
            f"Recusando sincronizacao: alvo nao parece localhost/auditoria "
            f"(host={parsed.hostname!r}, db={database!r})."
        )
    return psycopg2.connect(
        dbname=database,
        user=parsed.username,
        password=parsed.password,
        host=parsed.hostname,
        port=parsed.port or 5432,
    )


def load_official_operators() -> tuple[list[dict], set[str]]:
    workbook = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    sheet = workbook.active
    headers = [display_text(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    index = {name: pos for pos, name in enumerate(headers)}
    required = ["Matrícula", "Código Huawei", "Operadores", "Setor", "Função"]
    missing = [name for name in required if name not in index]
    if missing:
        raise SystemExit(f"Colunas ausentes na planilha oficial: {missing}")

    current_supervisor = ""
    supervisors: set[str] = set()
    by_name: dict[str, dict] = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        role = normalize_text(row[index["Função"]])
        name = display_text(row[index["Operadores"]])
        if not name:
            continue
        if "supervisor" in role:
            current_supervisor = name.title()
            supervisors.add(current_supervisor)
            continue
        if "operador" not in role:
            continue

        key = normalize_text(name)
        if not key or key in by_name:
            continue
        by_name[key] = {
            "nome": name,
            "nome_key": key,
            "matricula": clean_id(row[index["Matrícula"]]),
            "id_huawei": clean_id(row[index["Código Huawei"]]),
            "setor": display_text(row[index["Setor"]]),
            "escala": display_text(row[index["Setor"]]),
            "supervisor": current_supervisor,
            "manual": name.upper() in MANUAL_OPERATOR_NAMES,
        }

    return list(by_name.values()), supervisors


def dump_backup(cursor: RealDictCursor, backup_dir: Path) -> Path:
    backup_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = backup_dir / f"localhost_operadores_sync_backup_{stamp}.json"
    payload = {"created_at": datetime.now().isoformat(), "tables": {}}
    for table in ("colaboradores", "users"):
        cursor.execute(f"SELECT * FROM {table} ORDER BY id" if table == "colaboradores" else f"SELECT * FROM {table} ORDER BY username")
        payload["tables"][table] = cursor.fetchall()
    backup_path.write_text(json.dumps(payload, ensure_ascii=False, default=str, indent=2), encoding="utf-8")
    return backup_path


def pick_existing_id(existing: list[dict], operator: dict, claimed: set[int]) -> int | None:
    candidates: list[dict] = []
    name_key = operator["nome_key"]
    matricula = operator["matricula"]
    huawei = operator["id_huawei"]

    for row in existing:
        if int(row["id"]) in claimed:
            continue
        if normalize_text(row.get("nome")) == name_key:
            candidates.append(row)
    if not candidates and matricula:
        for row in existing:
            if int(row["id"]) not in claimed and clean_id(row.get("matricula")) == matricula:
                candidates.append(row)
    if not candidates and huawei:
        for row in existing:
            if int(row["id"]) in claimed:
                continue
            ids = {clean_id(row.get("id_huawei")), clean_id(row.get("id_telefonia"))}
            if huawei in ids:
                candidates.append(row)

    if not candidates:
        return None
    candidates.sort(key=lambda item: int(item["id"]))
    return int(candidates[0]["id"])


def apply_sync(cursor: RealDictCursor, operators: list[dict]) -> dict:
    cursor.execute("SELECT * FROM colaboradores ORDER BY id")
    existing = cursor.fetchall()
    claimed_ids: set[int] = set()
    inserted = 0
    updated = 0

    for operator in operators:
        existing_id = pick_existing_id(existing, operator, claimed_ids)
        params = {
            "nome": operator["nome"],
            "supervisor": operator["supervisor"],
            "setor": operator["setor"],
            "escala": operator["escala"],
            "status": "ATIVO",
            "matricula": operator["matricula"],
            "id_huawei": operator["id_huawei"],
            "id_telefonia": operator["id_huawei"],
            "auditavel": 1,
        }
        if existing_id is None:
            cursor.execute(
                """
                INSERT INTO colaboradores (
                    nome, supervisor, setor, escala, status, matricula,
                    id_huawei, id_telefonia, auditavel, atualizado_em
                )
                VALUES (
                    %(nome)s, %(supervisor)s, %(setor)s, %(escala)s, %(status)s,
                    %(matricula)s, %(id_huawei)s, %(id_telefonia)s, %(auditavel)s,
                    CURRENT_TIMESTAMP
                )
                RETURNING id
                """,
                params,
            )
            new_id = int(cursor.fetchone()["id"])
            claimed_ids.add(new_id)
            inserted += 1
        else:
            params["id"] = existing_id
            cursor.execute(
                """
                UPDATE colaboradores
                SET nome = %(nome)s,
                    supervisor = %(supervisor)s,
                    setor = %(setor)s,
                    escala = %(escala)s,
                    status = %(status)s,
                    matricula = %(matricula)s,
                    id_huawei = %(id_huawei)s,
                    id_telefonia = %(id_telefonia)s,
                    auditavel = %(auditavel)s,
                    softphone_number = '',
                    telefonia_account = '',
                    organizacao_telefonia = '',
                    tipo_agente = '',
                    status_telefonia = '',
                    atualizado_em = CURRENT_TIMESTAMP
                WHERE id = %(id)s
                """,
                params,
            )
            claimed_ids.add(existing_id)
            updated += 1

    cursor.execute(
        """
        SELECT c.id
        FROM colaboradores c
        LEFT JOIN audits a ON a.colaborador_id = c.id
        LEFT JOIN fechamento_cadeia_contatos f ON f.colaborador_id = c.id
        WHERE NOT (c.id = ANY(%s))
          AND a.id IS NULL
          AND f.id IS NULL
        """,
        (list(claimed_ids),),
    )
    removable_ids = [int(row["id"]) for row in cursor.fetchall()]
    deleted = 0
    if removable_ids:
        cursor.execute("DELETE FROM colaboradores WHERE id = ANY(%s)", (removable_ids,))
        deleted = cursor.rowcount

    cursor.execute(
        """
        UPDATE colaboradores
        SET status = 'INATIVO', auditavel = 0, atualizado_em = CURRENT_TIMESTAMP
        WHERE NOT (id = ANY(%s))
        """,
        (list(claimed_ids),),
    )
    inactivated = cursor.rowcount

    return {
        "official_unique_operators": len(operators),
        "updated": updated,
        "inserted": inserted,
        "deleted_unreferenced_extra": deleted,
        "inactivated_referenced_extra": inactivated,
    }


def validate(cursor: RealDictCursor) -> dict:
    cursor.execute("SELECT COUNT(*) AS total FROM colaboradores")
    raw_total = cursor.fetchone()["total"]
    cursor.execute("SELECT COUNT(*) AS total FROM colaboradores WHERE status = 'ATIVO' AND COALESCE(auditavel, 1) = 1")
    active_total = cursor.fetchone()["total"]
    cursor.execute("SELECT COUNT(DISTINCT lower(trim(nome))) AS total FROM colaboradores WHERE status = 'ATIVO' AND COALESCE(auditavel, 1) = 1")
    active_unique_names = cursor.fetchone()["total"]
    cursor.execute("SELECT COUNT(*) AS total FROM users WHERE role = 'supervisor'")
    supervisor_users = cursor.fetchone()["total"]
    return {
        "colaboradores_raw_total": raw_total,
        "active_auditable_total": active_total,
        "active_auditable_unique_names": active_unique_names,
        "supervisor_users": supervisor_users,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true", help="Aplica a sincronizacao. Sem isso roda dry-run.")
    parser.add_argument("--backup-dir", default=r"C:\backups_auditoria")
    args = parser.parse_args()

    database_url = load_env_database_url()
    operators, supervisors = load_official_operators()
    print(f"Planilha oficial: {len(operators)} operadores unicos, {len(supervisors)} supervisores.")

    conn = connect(database_url)
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cursor:
            cursor.execute("SELECT current_database() AS db, current_user AS user, inet_server_addr()::text AS addr, inet_server_port() AS port")
            target = cursor.fetchone()
            print(f"Banco alvo: db={target['db']} user={target['user']} addr={target['addr']} port={target['port']}")
            print("Antes:", validate(cursor))

            if not args.apply:
                conn.rollback()
                print("Dry-run concluido. Use --apply para aplicar.")
                return

            backup_path = dump_backup(cursor, Path(args.backup_dir))
            result = apply_sync(cursor, operators)
            after = validate(cursor)
            conn.commit()
            print(f"Backup: {backup_path}")
            print("Aplicacao:", result)
            print("Depois:", after)
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


if __name__ == "__main__":
    main()
