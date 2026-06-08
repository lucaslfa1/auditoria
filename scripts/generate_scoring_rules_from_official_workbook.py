from __future__ import annotations

from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any

import openpyxl
import yaml


ROOT = Path(__file__).resolve().parents[1]
SCORING_RULES_PATH = ROOT / "backend" / "db" / "scoring_rules.yaml"
OFFICIAL_WORKBOOK_PATH = (
    ROOT / "auditoria_criterios" / "criterios_pesos" / "CRITÉRIOS - PESOS -.xlsm"
)

ALERT_TO_SHEET = {
    "UTI-PRIORITARIO-MOT": "Prioritario",
    "UTI-PRIORITARIO-CLI": "C.Prioritario",
    "UTI-POSICAO-MOT": "Posição",
    "UTI-POSICAO-CLI": "C.Posição",
    "UTI-PARADA-MOT": "Parada",
    "UTI-PARADA-CLI": "C.Parada",
    "UTI-DESVIO-MOT": "Desvio",
    "UTI-DESVIO-CLI": "C.Desvio",
    "UTI-PONTO-APOIO": "Apoio",
    "UTI-PRIORITARIO-POLICIA": "Policial",
    "BAS-PRIORITARIO-POLICIA": "Policial",
    "CADASTRO-ANTECEDENTES": "Antecedente",
    "UNILEVER-DEVOLUCAO": "Devolução",
    "UNILEVER-CABINETS": "Cabinets",
    "UNILEVER-TRATATIVA": "Atuação",
    "UNILEVER-DISTRIBUICAO": "Distribuição",
    "UNILEVER-LOSSTREE": "Loss Tree",
    "LOGISTICA-ESTADIA": "Estadia",
    "LOGISTICA-TEMPERATURA-MOT": "Cont.Temp",
    "LOGISTICA-TEMPERATURA-CLI": "Cont.Temp .Clien",
    "LOGISTICA-DESLIG-TEMP-MOT": "Desl.Temp",
    "LOGISTICA-DESLIG-TEMP-CLI": "Desl.Temp.Clien",
    "LOGISTICA-ATRASO-ENTREGA": "Atraso",
    "LOGISTICA-PARADA": "Parada Indevida Logística",
    "LOGISTICA-DESVIO": "Desvio de Rota Logística ",
    "LOGISTICA-ATIVACAO-AE": "Ativação AE.Clien",
    "LOGISTICA-ATRASO": "Atr.Entrega.Clien",
    "LOGISTICA-POSICAO": "Posição em Atraso Logística",
    "LOGISTICA-TABORDA": "Taborda",
    "LOGISTICA-ATRASO-INICIO": "Atraso no Início de Viagem",
    "MONDELEZ-LOGISTICA-REVERSA": "Logística Reversa",
    "MONDELEZ-MONITORAMENTO-I": "Monitoramento I",
    "MONDELEZ-MONITORAMENTO-II": "Monitoramento II",
    "CHECKLIST-VEICULO": "Checklist",
    "CHECKLIST-RECEPTIVO": "Checklist",
    "RECEPTIVO-CHATBOT": "Receptivo",
    "CELULA-RECEPTIVO": "Receptivo",
}

MANUAL_LABEL_MARKERS = (
    "registrou corretamente o contato no sistema",
    "qualificação correta",
    "qualificação do atendimento",
    "anexou imagens",
    "registrado no sil",
    "encerrou o checklist no sil",
    "encerrou o atendimento no weon",
)


def _clean(value: Any) -> str:
    return " ".join(str(value or "").replace("\xa0", " ").split()).strip()


def _weight(value: Any) -> float:
    return float(Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def _split_weight(weight: float, parts: int) -> list[float]:
    cents = int((Decimal(str(weight)) * 100).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    base = cents // parts
    remainder = cents % parts
    return [float(Decimal(base + (1 if index < remainder else 0)) / Decimal(100)) for index in range(parts)]


def _is_split_identification(label: str) -> bool:
    normalized = _clean(label).lower()
    return (
        "saudação, nome, setor e empresa" in normalized
        or "saudação, nome, setor, empresa" in normalized
    )


def _description(importance: str, example: str) -> str:
    pieces = []
    if importance:
        pieces.append(importance)
    if example:
        pieces.append(f"Exemplo oficial: {example}")
    return " | ".join(pieces)


def _evaluation_type(label: str) -> str:
    normalized = _clean(label).lower()
    return "manual" if any(marker in normalized for marker in MANUAL_LABEL_MARKERS) else "auto"


def _official_rows(ws) -> list[dict[str, Any]]:
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        label = _clean(row[1] if len(row) > 1 else "")
        weight = row[2] if len(row) > 2 else None
        if not label or not isinstance(weight, (int, float)):
            continue
        rows.append(
            {
                "label": label,
                "weight": _weight(weight),
                "importance": _clean(row[3] if len(row) > 3 else ""),
                "example": _clean(row[4] if len(row) > 4 else ""),
            }
        )
    return rows


def _build_criteria(rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], int]:
    criteria: list[dict[str, Any]] = []
    split_count = 0
    for row in rows:
        desc = _description(row["importance"], row["example"])
        if _is_split_identification(row["label"]):
            split_count += 1
            official_desc = _description(
                f"Fracionamento do critério oficial: {row['label']}",
                row["example"],
            )
            for label, weight in zip(("Saudação?", "Nome?", "Setor/Empresa?"), _split_weight(row["weight"], 3)):
                criteria.append(
                    {
                        "label": label,
                        "weight": weight,
                        "deflator": 0.0,
                        "description": official_desc,
                    }
                )
            continue

        criterion = {
            "label": row["label"],
            "weight": row["weight"],
            "deflator": 0.0,
        }
        if desc:
            criterion["description"] = desc
        if _evaluation_type(row["label"]) == "manual":
            criterion["evaluation_type"] = "manual"
        criteria.append(criterion)
    return criteria, split_count


def main() -> None:
    with SCORING_RULES_PATH.open("r", encoding="utf-8") as file:
        scoring_rules = yaml.safe_load(file)

    workbook = openpyxl.load_workbook(OFFICIAL_WORKBOOK_PATH, data_only=True, read_only=True, keep_vba=False)

    total_criteria = 0
    total_splits = 0
    missing = []
    for alert in scoring_rules["alerts"]:
        alert_id = alert["id"]
        sheet_name = ALERT_TO_SHEET.get(alert_id)
        if not sheet_name or sheet_name not in workbook.sheetnames:
            missing.append(alert_id)
            continue

        ws = workbook[sheet_name]
        rows = _official_rows(ws)
        criteria, split_count = _build_criteria(rows)
        alert["context"] = _clean(ws.cell(1, 2).value)
        alert["criteria"] = criteria
        total_criteria += len(criteria)
        total_splits += split_count

    if missing:
        raise RuntimeError(f"Alertas sem aba oficial mapeada: {', '.join(missing)}")

    with SCORING_RULES_PATH.open("w", encoding="utf-8") as file:
        yaml.safe_dump(scoring_rules, file, allow_unicode=True, sort_keys=False, width=120)

    print(
        f"scoring_rules.yaml atualizado: {len(scoring_rules['alerts'])} alertas, "
        f"{total_criteria} critérios, {total_splits} critérios de identificação fracionados."
    )


if __name__ == "__main__":
    main()
